#!/usr/bin/env python3
"""
Excel Skill — Direct programmatic access to Excel files.
Layer 1 (Skill/API) — No UI automation needed.

Usage:
  python3 excel-skill.py read <filepath> [--sheet <name>] [--range <A1:B10>]
  python3 excel-skill.py read-cell <filepath> <cell> [--sheet <name>]
  python3 excel-skill.py write-cell <filepath> <cell> <value> [--sheet <name>]
  python3 excel-skill.py search <filepath> <query> [--sheet <name>]
  python3 excel-skill.py list-sheets <filepath>
  python3 excel-skill.py info <filepath>

All output is JSON for easy parsing by the agent.
"""
import sys, json, argparse, os

def main():
    parser = argparse.ArgumentParser()
    sub = parser.add_subparsers(dest='command')

    # read: dump sheet or range as JSON array of rows
    p_read = sub.add_parser('read')
    p_read.add_argument('filepath')
    p_read.add_argument('--sheet', default=None)
    p_read.add_argument('--range', default=None)
    p_read.add_argument('--max-rows', type=int, default=100,
                        help='Max data rows to return (default 100, use 0 for unlimited)')

    # read-cell: single cell value
    p_rc = sub.add_parser('read-cell')
    p_rc.add_argument('filepath')
    p_rc.add_argument('cell')  # e.g. "B3"
    p_rc.add_argument('--sheet', default=None)

    # write-cell: set a cell value
    p_wc = sub.add_parser('write-cell')
    p_wc.add_argument('filepath')
    p_wc.add_argument('cell')
    p_wc.add_argument('value')
    p_wc.add_argument('--sheet', default=None)

    # search: find rows containing a value
    p_search = sub.add_parser('search')
    p_search.add_argument('filepath')
    p_search.add_argument('query')
    p_search.add_argument('--sheet', default=None)

    # list-sheets
    p_ls = sub.add_parser('list-sheets')
    p_ls.add_argument('filepath')

    # info: sheet names, dimensions, headers
    p_info = sub.add_parser('info')
    p_info.add_argument('filepath')

    args = parser.parse_args()
    if not args.command:
        parser.print_help()
        sys.exit(1)

    try:
        import openpyxl
    except ImportError:
        print(json.dumps({"error": "openpyxl not installed. Run: pip3 install openpyxl"}))
        sys.exit(1)

    filepath = os.path.expanduser(args.filepath)
    if not os.path.exists(filepath):
        print(json.dumps({"error": f"File not found: {filepath}"}))
        sys.exit(1)

    try:
        if args.command == 'info':
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            sheets = []
            for name in wb.sheetnames:
                ws = wb[name]
                # Read first row as headers
                headers = []
                for cell in next(ws.iter_rows(min_row=1, max_row=1, values_only=False), []):
                    headers.append(str(cell.value) if cell.value is not None else "")
                sheets.append({
                    "name": name,
                    "rows": ws.max_row,
                    "cols": ws.max_column,
                    "headers": headers
                })
            wb.close()
            print(json.dumps({"file": filepath, "sheets": sheets}))

        elif args.command == 'list-sheets':
            wb = openpyxl.load_workbook(filepath, read_only=True)
            print(json.dumps({"sheets": wb.sheetnames}))
            wb.close()

        elif args.command == 'read':
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb[args.sheet] if args.sheet else wb.active
            rows = []
            if args.range:
                for row in ws[args.range]:
                    rows.append([str(c.value) if c.value is not None else "" for c in row])
            else:
                for row in ws.iter_rows(values_only=True):
                    rows.append([str(v) if v is not None else "" for v in row])
            wb.close()
            if rows:
                headers = rows[0]
                data_rows = rows[1:]
                total = len(data_rows)
                max_rows = args.max_rows
                truncated = False
                if max_rows > 0 and total > max_rows:
                    data_rows = data_rows[:max_rows]
                    truncated = True
                data = [dict(zip(headers, r)) for r in data_rows]
                result = {"headers": headers, "data": data, "rowCount": len(data), "totalRows": total}
                if truncated:
                    result["truncated"] = True
                    result["note"] = f"Showing {max_rows} of {total} rows. Use --max-rows 0 for all, or --range A1:Z{max_rows+1} for a specific range."
                print(json.dumps(result))
            else:
                print(json.dumps({"headers": [], "data": [], "rowCount": 0, "totalRows": 0}))

        elif args.command == 'read-cell':
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb[args.sheet] if args.sheet else wb.active
            val = ws[args.cell].value
            wb.close()
            print(json.dumps({"cell": args.cell, "value": str(val) if val is not None else None}))

        elif args.command == 'write-cell':
            wb = openpyxl.load_workbook(filepath)
            ws = wb[args.sheet] if args.sheet else wb.active
            # Try to convert to number
            try:
                val = float(args.value)
                if val == int(val):
                    val = int(val)
            except ValueError:
                val = args.value
            ws[args.cell] = val
            wb.save(filepath)
            wb.close()
            print(json.dumps({"cell": args.cell, "value": str(val), "saved": True}))

        elif args.command == 'search':
            wb = openpyxl.load_workbook(filepath, read_only=True, data_only=True)
            ws = wb[args.sheet] if args.sheet else wb.active
            query = args.query.lower()
            results = []
            headers = None
            for i, row in enumerate(ws.iter_rows(values_only=True), 1):
                row_strs = [str(v) if v is not None else "" for v in row]
                if i == 1:
                    headers = row_strs
                    continue
                if any(query in cell.lower() for cell in row_strs):
                    results.append({"row": i, "data": dict(zip(headers, row_strs))})
            wb.close()
            print(json.dumps({"query": args.query, "matches": results, "matchCount": len(results)}))

    except Exception as e:
        print(json.dumps({"error": str(e)}))
        sys.exit(1)

if __name__ == '__main__':
    main()
